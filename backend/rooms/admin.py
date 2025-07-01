import csv
from django.contrib import admin
from django.utils.timezone import now, make_aware
from django.utils.safestring import mark_safe
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import render
from django.urls import path
from .models import Room, Guest, Reservation, Building, Floor
from .forms import ReservationAdminForm, ImportReservationForm, RoomAdminForm  # Add RoomAdminForm here
from django.db import transaction
from .utils import parse_date
import openpyxl
from datetime import datetime
from dal import autocomplete
import json

class FloorAutocomplete(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        qs = Floor.objects.all()
        building_id = self.forwarded.get('building', None)
        if building_id:
            qs = qs.filter(building_id=building_id)
        return qs

@admin.register(Building)
class BuildingAdmin(admin.ModelAdmin):
    list_display = ['name', 'description']
    search_fields = ['name']

@admin.register(Floor)
class FloorAdmin(admin.ModelAdmin):
    list_display = ['building', 'number', 'name']
    list_filter = ['building']
    search_fields = ['building__name', 'number', 'name']
    ordering = ['building', 'number']

@admin.register(Room)
class RoomAdmin(admin.ModelAdmin):
    form = RoomAdminForm
    list_display = [
        'number',
        'get_building_info',
        'room_type',
        'capacity',
        'has_fan',
        'get_availability_status',
        'has_attached_bathroom',
        'has_kitchen',
        'donation_due',
        'needs_repair',
        'needs_supplies',
        'contents_summary'
    ]
    actions = ['mark_as_cleaned']
    list_display_links = ['number']
    list_filter = [
        'building',
        'room_type',
        'has_fan',
        'has_attached_bathroom',
        'has_kitchen',
        'needs_repair',
        'needs_supplies',
        'is_available'
    ]
    
    search_fields = ['number', 'building__name', 'floor__number']
    fieldsets = (
        ('Basic Information', {
            'fields': ('number', 'building', 'floor', 'room_type', 'capacity', 'donation_due', 'is_available')
        }),
        ('Facilities', {
            'fields': ('has_fan', 'has_attached_bathroom', 'has_kitchen')
        }),
        ('Maintenance', {
            'fields': ('needs_repair', 'repair_notes', 'needs_supplies', 'supply_notes')
        }),
        ('Room Contents', {
            'fields': ('beds', 'pillows', 'mats', 'foldable_cots', 'chairs', 'stools')
        }),
        ('Notes', {
            'fields': ('notes',)
        }),
    )

    class Media:
        js = ('admin/room_admin.js',)

    def changeform_view(self, request, object_id=None, form_url='', extra_context=None):
        extra_context = extra_context or {}
        # Build a dict: floor_id -> allowed_room_types
        allowed_types = {str(f.id): f.allowed_room_types for f in Floor.objects.all()}
        extra_context['allowed_room_types_json'] = mark_safe(json.dumps(allowed_types))
        return super().changeform_view(request, object_id, form_url, extra_context=extra_context)
    def contents_summary(self, obj):
        """Admin wrapper for the model's get_contents_summary method"""
        return obj.get_contents_summary()
    contents_summary.short_description = "Room Contents"

    @admin.action(description="Mark selected rooms as cleaned")
    def mark_as_cleaned(self, request, queryset):
        updated = queryset.update(needs_cleaning=False)
        self.message_user(request, f"{updated} room(s) marked as cleaned")

    def get_availability_status(self, obj):
        return obj.get_availability_status()
    get_availability_status.short_description = 'Availability'

    def get_building_info(self, obj):
        return f"{obj.building.name if obj.building else ''} - Floor {obj.floor.number if obj.floor else ''}"
    get_building_info.short_description = 'Building - Floor'
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related('building', 'floor').prefetch_related('reservation_set')

@admin.register(Guest)
class GuestAdmin(admin.ModelAdmin):
    list_display = ['full_name', 'phone_number', 'email', 'country', 'city']
    search_fields = ['full_name', 'phone_number', 'email', 'country', 'city']

@admin.register(Reservation)
class ReservationAdmin(admin.ModelAdmin):
    form = ReservationAdminForm
    class Media:
        js = ('admin/js/jquery.init.js',)
    list_display = (
        "id", "room", "guest_list", "check_in_date", "check_out_date",
        "is_checked_in", "is_checked_out", "is_paid", "is_cancelled", "allow_guest_overlap"
    )
    fieldsets = (
        ('Basic Info', {'fields': ('room', 'guests', 'check_in_date', 'check_out_date')}),
    )
    filter_horizontal = ("guests",)
    list_filter = ['is_checked_in', 'is_checked_out', 'is_paid', 'is_cancelled', 'allow_guest_overlap']
    search_fields = ['guests__full_name', 'room__number']
    actions = ['mark_as_checked_in', 'mark_as_checked_out', 'cancel_reservation', 'export_as_csv', 'export_as_excel']
    change_list_template = 'admin/reservations_change_list.html'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-reservations/', self.import_reservations, name='import_reservations'),
            path('export-csv/', self.export_all_csv, name='export_reservations_csv'),
            path('export-excel/', self.export_all_excel, name='export_reservations_excel'),
        ]
        return custom_urls + urls

    def guest_list(self, obj):
        return ", ".join([g.full_name for g in obj.guests.all()])
    guest_list.short_description = 'Guests'

    def get_room_info(self, obj):
        if obj.room:
            building = obj.room.building.name if obj.room.building else 'No Building'
            floor = obj.room.floor.number if obj.room.floor else 'No Floor'
            return f"#{obj.room.number} ({building} - Floor {floor})"
        return "No Room Assigned"
    get_room_info.short_description = 'Room'
    get_room_info.admin_order_field = 'room__number'

    @admin.action(description="Mark selected reservations as Checked In")
    def mark_as_checked_in(self, request, queryset):
        for reservation in queryset:
            reservation.is_checked_in = True
            reservation.is_cancelled = False
            reservation.is_checked_out = False
            try:
                reservation.save()
            except Exception as e:
                self.message_user(request, f"Error updating reservation {reservation.id}: {str(e)}", level=messages.ERROR)
        self.message_user(request, f"{queryset.count()} reservation(s) marked as checked in.")

    @admin.action(description="Mark selected reservations as Checked Out")
    def mark_as_checked_out(self, request, queryset):
        updated = 0
        for reservation in queryset:
            reservation.is_checked_out = True
            try:
                reservation.save()
                updated += 1
            except Exception as e:
                self.message_user(request, f"Error updating reservation {reservation.id}: {str(e)}", level=messages.ERROR)
        self.message_user(request, f"{updated} reservation(s) marked as checked out.")

    @admin.action(description="Cancel selected reservations")
    def cancel_reservation(self, request, queryset):
        updated = 0
        for reservation in queryset:
            reservation.is_cancelled = True
            try:
                reservation.save()
                updated += 1
            except Exception as e:
                self.message_user(request, f"Error cancelling reservation {reservation.id}: {str(e)}", level=messages.ERROR)
        self.message_user(request, f"{updated} reservation(s) cancelled.")

    @admin.action(description="Export selected reservations as CSV")
    def export_as_csv(self, request, queryset):
        meta = self.model._meta
        field_names = [field.name for field in meta.fields]
        field_names += ['guests_list', 'room_info']

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename={meta}_reservations_{datetime.now().date()}.csv'
        
        writer = csv.writer(response)
        writer.writerow(field_names)
        
        for obj in queryset:
            row = []
            for field in field_names:
                if field == 'guests_list':
                    value = ", ".join([g.full_name for g in obj.guests.all()])
                elif field == 'room_info':
                    value = f"{obj.room.number if obj.room else 'No room'} ({obj.room.building.name if obj.room and obj.room.building else ''})"
                else:
                    value = getattr(obj, field)
                    if isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M')
                row.append(value)
            writer.writerow(row)
        
        return response
    export_as_csv.short_description = "Export Selected as CSV"

    @admin.action(description="Export selected reservations as Excel")
    def export_as_excel(self, request, queryset):
        meta = self.model._meta
        field_names = [field.name for field in meta.fields] 
        field_names += ['guests_list', 'room_info']

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={meta}_reservations_{datetime.now().date()}.xlsx'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reservations"

        for col_num, field_name in enumerate(field_names, 1):
            ws.cell(row=1, column=col_num, value=field_name)

        for row_num, obj in enumerate(queryset, 2):
            for col_num, field_name in enumerate(field_names, 1):
                if field_name == 'guests_list':
                    value = ", ".join([g.full_name for g in obj.guests.all()])
                elif field_name == 'room_info':
                    value = f"{obj.room.number if obj.room else 'No room'} ({obj.room.building.name if obj.room and obj.room.building else ''})"
                else:
                    value = getattr(obj, field_name)
                    if isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M')
                ws.cell(row=row_num, column=col_num, value=value)

        wb.save(response)
        return response
    export_as_excel.short_description = "Export Selected as Excel"

    def export_all_csv(self, request):
        return self.export_as_csv(request, self.get_queryset(request))

    def export_all_excel(self, request):
        return self.export_as_excel(request, self.get_queryset(request))

    def import_reservations(self, request):
        if request.method == 'POST':
            form = ImportReservationForm(request.POST, request.FILES)
            if form.is_valid():
                try:
                    csv_file = request.FILES['csv_file']
                    decoded_file = csv_file.read().decode('utf-8').splitlines()
                    reader = csv.DictReader(decoded_file)
                    
                    created_count = 0
                    error_messages = []
                    
                    with transaction.atomic():
                        for row in reader:
                            try:
                                room_number = row.get('room_number')
                                room = Room.objects.filter(number=room_number).first() if room_number else None
                                
                                reservation = Reservation(
                                    room=room,
                                    check_in_date=make_aware(parse_date(row['check_in_date'])),
                                    check_out_date=make_aware(parse_date(row['check_out_date'])),
                                    is_checked_in=row.get('is_checked_in', '').lower() == 'true',
                                    is_checked_out=row.get('is_checked_out', '').lower() == 'true',
                                    is_paid=row.get('is_paid', '').lower() == 'true',
                                    is_cancelled=row.get('is_cancelled', '').lower() == 'true',
                                    allow_guest_overlap=row.get('allow_guest_overlap', '').lower() == 'true',
                                    notes=row.get('notes', '')
                                )
                                reservation.save()
                                
                                guest_names = [name.strip() for name in row.get('guests', '').split(',') if name.strip()]
                                for name in guest_names:
                                    guest, created = Guest.objects.get_or_create(full_name=name)
                                    reservation.guests.add(guest)
                                
                                created_count += 1
                            except Exception as e:
                                error_messages.append(f"Error processing row {reader.line_num}: {str(e)}")
                    
                    if created_count > 0:
                        self.message_user(request, f"Successfully imported {created_count} reservations.")
                    if error_messages:
                        for msg in error_messages:
                            self.message_user(request, msg, level=messages.ERROR)
                    return HttpResponseRedirect("../")
                
                except Exception as e:
                    self.message_user(request, f"Error importing file: {str(e)}", level=messages.ERROR)
        else:
            form = ImportReservationForm()
        
        context = {
            'form': form,
            'opts': self.model._meta,
            'title': 'Import Reservations',
        }
        return render(request, 'admin/import_reservations.html', context)

    def save_model(self, request, obj, form, change):
        try:
            super().save_model(request, obj, form, change)
            form.save_m2m()
            
            if hasattr(obj, 'check_guest_overlaps'):
                conflicts = obj.check_guest_overlaps()
                if conflicts and not obj.allow_guest_overlap:
                    conflict_messages = []
                    for conflict in conflicts:
                        guest = conflict['guest']
                        for res in conflict['reservations']:
                            conflict_messages.append(
                                f"{guest.full_name} has overlapping reservation #{res.id} "
                                f"({res.check_in_date} to {res.check_out_date})"
                            )
                    
                    messages.warning(
                        request,
                        "Guest overlap detected:\n" + "\n".join(conflict_messages),
                        extra_tags='guest_conflict'
                    )

            if hasattr(obj, 'room') and obj.room:
                self.update_room_availability(obj)
                
        except Exception as e:
            messages.error(request, f"Error saving reservation: {str(e)}")
            raise

    def save_related(self, request, form, formsets, change):
        try:
            for formset in formsets:
                self.save_formset(request, form, formset, change=change)
        except Exception as e:
            messages.error(request, f"Error saving related objects: {str(e)}")
            raise

    def update_room_availability(self, obj):
        try:
            if not hasattr(obj, 'room') or not obj.room:
                return
                
            if obj.is_cancelled or obj.is_checked_out:
                obj.room.is_available = True
                if obj.is_checked_out:
                    obj.room.needs_cleaning = True
            elif obj.is_checked_in:
                active_reservations = Reservation.objects.filter(
                    room=obj.room,
                    is_checked_in=True,
                    is_checked_out=False,
                    is_cancelled=False
                ).exclude(pk=obj.pk)

                total_checked_in_guests = sum(r.guests.count() for r in active_reservations) + obj.guests.count()
                obj.room.is_available = total_checked_in_guests < obj.room.capacity
                obj.room.needs_cleaning = False
            
            obj.room.save()
        except Exception as e:
            raise ValidationError(f"Error updating room availability: {str(e)}")